"""One-time data loading: Excel workbook -> SQLite tables + Chroma collections.

Principle: Separation of Concerns — loading/ETL lives here and nowhere else.
Runtime query code (sqlite_store, tools) never parses Excel.

Principle: Idempotency — loading checks whether work is already done, so the
app is safe to start repeatedly and "works out of the box" on first run.
"""
from __future__ import annotations

import sqlite3
from pathlib import Path

import openpyxl

from app.data.interfaces import StructuredStore, VectorStore

# Sheets we mirror verbatim into SQLite tables.
_SHEETS = [
    "alarm_reference",
    "current_incidents",
    "engineer_directory",
    "engineer_feedback",
    "equipment_master",
    "escalation_rules",
    "incident_history",
    "lot_wip",
    "maintenance_records",
    "sensor_readings",
    "sop_knowledge_base",
    "test_cases",
]


def _is_already_loaded(conn: sqlite3.Connection) -> bool:
    """True only if a sentinel table is present.

    Defensive: checking for real tables (not just that a file exists) means a
    half-written or empty .db left behind by a crash is detected and rebuilt,
    rather than silently skipped. This keeps first-run behaviour reliable.
    """
    row = conn.execute(
        "SELECT name FROM sqlite_master WHERE type='table' AND name='equipment_master'"
    ).fetchone()
    return row is not None


def load_excel_to_sqlite(dataset_path: str, sqlite_path: str) -> None:
    """Create one table per sheet. Idempotent: rebuilds only if not fully loaded."""
    db_file = Path(sqlite_path)
    db_file.parent.mkdir(parents=True, exist_ok=True)

    conn = sqlite3.connect(sqlite_path)
    try:
        if _is_already_loaded(conn):
            return  # already loaded — nothing to do

        wb = openpyxl.load_workbook(dataset_path, data_only=True)
        for sheet in _SHEETS:
            if sheet not in wb.sheetnames:
                continue
            ws = wb[sheet]
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                continue
            headers = [str(h) for h in rows[0]]
            cols = ", ".join(f'"{h}" TEXT' for h in headers)
            # DROP first so a partial previous load can't leave stale tables.
            conn.execute(f'DROP TABLE IF EXISTS "{sheet}"')
            conn.execute(f'CREATE TABLE "{sheet}" ({cols})')
            placeholders = ", ".join("?" for _ in headers)
            conn.executemany(
                f'INSERT INTO "{sheet}" VALUES ({placeholders})',
                [tuple("" if v is None else str(v) for v in r) for r in rows[1:]],
            )
        conn.commit()
    finally:
        conn.close()


def build_vector_index(store: StructuredStore, vectors: VectorStore) -> None:
    """Embed historical incidents and SOPs for semantic retrieval.

    Idempotent: skips collections that are already populated.
    """
    # --- similar-incidents collection --------------------------------------
    if vectors.count("incidents") == 0:
        history = store.get_incident_history()
        if history:
            ids, docs, metas = [], [], []
            for h in history:
                # A compact natural-language description is what we embed, so
                # free-text queries match on meaning, not exact column values.
                doc = (
                    f"{h.get('alarm_code')} on {h.get('equipment_id')}: "
                    f"root cause {h.get('root_cause')}. "
                    f"Corrective action: {h.get('corrective_action')}. "
                    f"Downtime {h.get('downtime_minutes')} min, "
                    f"impact {h.get('product_impact')}."
                )
                ids.append(str(h.get("incident_id")))
                docs.append(doc)
                metas.append(
                    {
                        "incident_id": str(h.get("incident_id")),
                        "equipment_id": str(h.get("equipment_id")),
                        "alarm_code": str(h.get("alarm_code")),
                        "root_cause": str(h.get("root_cause")),
                        "corrective_action": str(h.get("corrective_action")),
                        "downtime_minutes": str(h.get("downtime_minutes")),
                        "closure_status": str(h.get("closure_status")),
                    }
                )
            vectors.upsert("incidents", ids, docs, metas)

    # --- SOP collection ----------------------------------------------------
    if vectors.count("sops") == 0:
        sops = store.get_sops()
        if sops:
            ids, docs, metas = [], [], []
            for s in sops:
                doc = (
                    f"{s.get('title')} ({s.get('alarm_code')}, "
                    f"{s.get('tool_type')}): {s.get('troubleshooting_steps')}"
                )
                ids.append(str(s.get("sop_id")))
                docs.append(doc)
                metas.append(
                    {
                        "sop_id": str(s.get("sop_id")),
                        "alarm_code": str(s.get("alarm_code")),
                        "tool_type": str(s.get("tool_type")),
                        "title": str(s.get("title")),
                        "troubleshooting_steps": str(s.get("troubleshooting_steps")),
                    }
                )
            vectors.upsert("sops", ids, docs, metas)
